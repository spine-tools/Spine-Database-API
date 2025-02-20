######################################################################################################################
# Copyright (C) 2017-2022 Spine project consortium
# Copyright Spine Database API contributors
# This file is part of Spine Database API.
# Spine Database API is free software: you can redistribute it and/or modify it under the terms of the GNU Lesser
# General Public License as published by the Free Software Foundation, either version 3 of the License, or (at your
# option) any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
# without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Lesser General
# Public License for more details. You should have received a copy of the GNU Lesser General Public License along with
# this program. If not, see <http://www.gnu.org/licenses/>.
######################################################################################################################

import pathlib
import pickle
import tempfile
import unittest
import openpyxl
from spinedb_api.exception import ReaderError
from spinedb_api.spine_io.importers.excel_reader import ExcelReader
from spinedb_api.spine_io.importers.reader import TableProperties


class TestExcelReader(unittest.TestCase):
    def test_get_tables_and_properties(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = pathlib.Path(temp_dir) / "test.xlsx"
            workbook = openpyxl.Workbook()
            workbook.save(file_path)
            reader = ExcelReader(None)
            reader.connect_to_source(str(file_path))
            try:
                tables = reader.get_tables_and_properties()
            finally:
                reader.disconnect()
            self.assertEqual(tables, {"Sheet": TableProperties()})

    def test_get_data(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = pathlib.Path(temp_dir) / "test.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append([11, 12, 13])
            sheet.append([21, 22, 23])
            workbook.save(file_path)
            reader = ExcelReader(None)
            reader.connect_to_source(str(file_path))
            try:
                tables = reader.get_tables_and_properties()
                self.assertEqual(len(tables), 1)
                table_name = next(iter(tables))
                self.assertEqual(table_name, "Sheet")
                data, headers = reader.get_data(table_name, tables[table_name].options)
            finally:
                reader.disconnect()
            self.assertEqual(headers, [])
            self.assertEqual(data, [[11, 12, 13], [21, 22, 23]])

    def test_get_data_with_options(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = pathlib.Path(temp_dir) / "test.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append([1, 2, 3, None, 5])
            sheet.append(["header 1", "header 2", "header 3", None, "header 5"])
            sheet.append([11, 12, 13, None, 15])
            sheet.append([21, 22, 23, None, 25])
            sheet.append([])
            sheet.append([41, 42, 43, None, 45])
            workbook.save(file_path)
            reader = ExcelReader(None)
            reader.connect_to_source(str(file_path))
            try:
                tables = reader.get_tables_and_properties()
                self.assertEqual(len(tables), 1)
                table_name, properties = next(iter(tables.items()))
                self.assertEqual(table_name, "Sheet")
                options = properties.options
                options["header"] = True
                options["row"] = 1
                options["column"] = 1
                options["read_until_col"] = True
                options["read_until_row"] = True
                data, headers = reader.get_data(table_name, tables[table_name].options)
            finally:
                reader.disconnect()
            self.assertEqual(headers, ["header 2", "header 3"])
            self.assertEqual(data, [[12, 13], [22, 23]])

    def test_get_table_cell(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = pathlib.Path(temp_dir) / "test.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append([11, 12, 13])
            sheet.append([21, 22, 23])
            workbook.save(file_path)
            reader = ExcelReader(None)
            reader.connect_to_source(str(file_path))
            try:
                data = reader.get_table_cell("Sheet", 1, 2, {})
            finally:
                reader.disconnect()
            self.assertEqual(data, 23)

    def test_get_table_cell_with_options(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = pathlib.Path(temp_dir) / "test.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append([1, 2, 3, None, 5])
            sheet.append(["header 1", "header 2", "header 3", None, "header 5"])
            sheet.append([11, 12, 13, None, 15])
            sheet.append([21, 22, 23, None, 25])
            sheet.append([])
            sheet.append([41, 42, 43, None, 45])
            workbook.save(file_path)
            reader = ExcelReader(None)
            reader.connect_to_source(str(file_path))
            try:
                tables = reader.get_tables_and_properties()
                self.assertEqual(len(tables), 1)
                table_name, properties = next(iter(tables.items()))
                self.assertEqual(table_name, "Sheet")
                options = properties.options
                options["header"] = True
                options["row"] = 1
                options["column"] = 1
                data = reader.get_table_cell(table_name, 0, 1, tables[table_name].options)
            finally:
                reader.disconnect()
            self.assertEqual(data, 13)

    def test_get_table_cell_raises_when_row_table_is_not_in_workbook(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = pathlib.Path(temp_dir) / "test.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append([11, 12, 13])
            sheet.append([21, 22, 23])
            workbook.save(file_path)
            reader = ExcelReader(None)
            reader.connect_to_source(str(file_path))
            try:
                tables = reader.get_tables_and_properties()
                self.assertEqual(len(tables), 1)
                table_name = next(iter(tables))
                self.assertEqual(table_name, "Sheet")
                with self.assertRaisesRegex(ReaderError, "no sheet called 'No such sheet'"):
                    reader.get_table_cell("No such sheet", 0, 0, tables[table_name].options)
            finally:
                reader.disconnect()

    def test_reader_is_picklable(self):
        reader = ExcelReader(None)
        pickled = pickle.dumps(reader)
        self.assertTrue(pickled)


if __name__ == "__main__":
    unittest.main()
