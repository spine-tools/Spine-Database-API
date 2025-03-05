######################################################################################################################
# Copyright (C) 2017-2022 Spine project consortium
# Copyright Spine Database API contributors
# This file is part of Spine Database API.
# Spine Database API is free software: you can redistribute it and/or modify it under the terms of the GNU Lesser
# General Public License as published by the Free Software Foundation, either version 3 of the License, or (at your
# option) any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
# without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Lesser General
# Public License for more details. You should have received a copy of the GNU Lesser General Public License along with
# this program. If not, see <http://www.gnu.org/licenses/>.
######################################################################################################################
""" Unit tests for Excel writer. """
from itertools import zip_longest
import os.path
from tempfile import TemporaryDirectory
import unittest
from openpyxl import load_workbook
from spinedb_api import DatabaseMapping, Map, import_object_classes, import_objects
from spinedb_api.export_mapping import entity_export, entity_parameter_value_export
from spinedb_api.mapping import Position
from spinedb_api.spine_io.exporters.excel_writer import ExcelWriter
from spinedb_api.spine_io.exporters.writer import write
from tests.mock_helpers import AssertSuccessTestCase


class TestExcelWriter(AssertSuccessTestCase):
    def test_write_empty_database(self):
        with TemporaryDirectory() as temp_dir:
            with DatabaseMapping("sqlite://", create=True) as db_map:
                root_mapping = entity_export(0, 1)
                path = os.path.join(temp_dir, "test.xlsx")
                writer = ExcelWriter(path)
                write(db_map, writer, root_mapping)
                workbook = load_workbook(path, read_only=True)
                self.assertEqual(workbook.sheetnames, ["Sheet1"])
                sheet = workbook["Sheet1"]
                self.assertEqual(sheet.calculate_dimension(), "A1:A1")
                workbook.close()

    def test_write_single_object_class_and_object(self):
        with TemporaryDirectory() as temp_dir:
            with DatabaseMapping("sqlite://", create=True) as db_map:
                self._assert_imports(import_object_classes(db_map, ("oc",)))
                self._assert_imports(import_objects(db_map, (("oc", "o1"),)))
                db_map.commit_session("Add test data.")
                root_mapping = entity_export(0, 1)
                path = os.path.join(temp_dir, "test.xlsx")
                writer = ExcelWriter(path)
                write(db_map, writer, root_mapping)
                workbook = load_workbook(path, read_only=True)
                self.assertEqual(workbook.sheetnames, ["Sheet1"])
                expected = [["oc", "o1"]]
                self.check_sheet(workbook, "Sheet1", expected)
                workbook.close()

    def test_write_to_existing_sheet(self):
        with TemporaryDirectory() as temp_dir:
            with DatabaseMapping("sqlite://", create=True) as db_map:
                self._assert_imports(import_object_classes(db_map, ("Sheet1",)))
                self._assert_imports(import_objects(db_map, (("Sheet1", "o1"), ("Sheet1", "o2"))))
                db_map.commit_session("Add test data.")
                root_mapping = entity_export(Position.table_name, 0)
                path = os.path.join(temp_dir, "test.xlsx")
                writer = ExcelWriter(path)
                write(db_map, writer, root_mapping)
                workbook = load_workbook(path, read_only=True)
                self.assertEqual(workbook.sheetnames, ["Sheet1"])
                expected = [["o1"], ["o2"]]
                self.check_sheet(workbook, "Sheet1", expected)
                workbook.close()

    def test_write_to_named_sheets(self):
        with TemporaryDirectory() as temp_dir:
            with DatabaseMapping("sqlite://", create=True) as db_map:
                self._assert_imports(import_object_classes(db_map, ("oc1", ("oc2"))))
                self._assert_imports(import_objects(db_map, (("oc1", "o11"), ("oc1", "o12"), ("oc2", "o21"))))
                db_map.commit_session("Add test data.")
                root_mapping = entity_export(Position.table_name, 1)
                path = os.path.join(temp_dir, "test.xlsx")
                writer = ExcelWriter(path)
                write(db_map, writer, root_mapping)
                workbook = load_workbook(path, read_only=True)
                self.assertEqual(workbook.sheetnames, ["oc1", "oc2"])
                expected = [[None, "o11"], [None, "o12"]]
                self.check_sheet(workbook, "oc1", expected)
                expected = [[None, "o21"]]
                self.check_sheet(workbook, "oc2", expected)
                workbook.close()

    def test_append_to_anonymous_table(self):
        with TemporaryDirectory() as temp_dir:
            with DatabaseMapping("sqlite://", create=True) as db_map:
                self._assert_imports(import_object_classes(db_map, ("oc",)))
                self._assert_imports(import_objects(db_map, (("oc", "o1"),)))
                db_map.commit_session("Add test data.")
                root_mapping1 = entity_export(0, 1)
                root_mapping2 = entity_export(0, 1)
                path = os.path.join(temp_dir, "test.xlsx")
                writer = ExcelWriter(path)
                write(db_map, writer, root_mapping1, root_mapping2)
                workbook = load_workbook(path, read_only=True)
                self.assertEqual(workbook.sheetnames, ["Sheet1"])
                expected = [["oc", "o1"], ["oc", "o1"]]
                self.check_sheet(workbook, "Sheet1", expected)
                workbook.close()

    def test_append_to_named_table(self):
        with TemporaryDirectory() as temp_dir:
            with DatabaseMapping("sqlite://", create=True) as db_map:
                self._assert_imports(import_object_classes(db_map, ("oc",)))
                self._assert_imports(import_objects(db_map, (("oc", "o1"),)))
                db_map.commit_session("Add test data.")
                root_mapping1 = entity_export(Position.table_name, 0)
                root_mapping2 = entity_export(Position.table_name, 0)
                path = os.path.join(temp_dir, "test.xlsx")
                writer = ExcelWriter(path)
                write(db_map, writer, root_mapping1, root_mapping2)
                workbook = load_workbook(path, read_only=True)
                self.assertEqual(workbook.sheetnames, ["oc"])
                expected = [["o1"], ["o1"]]
                self.check_sheet(workbook, "oc", expected)
                workbook.close()

    def test_value_indexes_remain_unsorted(self):
        with TemporaryDirectory() as temp_dir:
            with DatabaseMapping("sqlite://", create=True) as db_map:
                db_map.add_entity_class(name="Sheet1")
                db_map.add_parameter_definition(entity_class_name="Sheet1", name="z")
                db_map.add_entity(entity_class_name="Sheet1", name="o1")
                db_map.add_parameter_value(
                    entity_class_name="Sheet1",
                    entity_byname=("o1",),
                    parameter_definition_name="z",
                    alternative_name="Base",
                    parsed_value=Map(["T02", "T01"], [1.1, 1.2]),
                )
                db_map.commit_session("Add test data.")
                root_mapping = entity_parameter_value_export(
                    Position.table_name, 0, Position.hidden, 1, None, None, 2, 3, 6, [4], [5]
                )
                path = os.path.join(temp_dir, "test.xlsx")
                writer = ExcelWriter(path)
                write(db_map, writer, root_mapping)
                workbook = load_workbook(path, read_only=True)
                try:
                    self.assertEqual(workbook.sheetnames, ["Sheet1"])
                    expected = [
                        ["z", "o1", "Base", "1d_map", "x", "T02", 1.1],
                        ["z", "o1", "Base", "1d_map", "x", "T01", 1.2],
                    ]
                    self.check_sheet(workbook, "Sheet1", expected)
                finally:
                    workbook.close()

    def check_sheet(self, workbook, sheet_name, expected):
        """
        Args:
            workbook (Workbook): a workbook to check
            sheet_name (str): sheet name
            expected (list): expected rows
        """
        sheet = workbook[sheet_name]
        for row, expected_row in zip_longest(sheet.iter_rows(), expected):
            values = [cell.value for cell in row]
            self.assertEqual(values, expected_row)


if __name__ == "__main__":
    unittest.main()
