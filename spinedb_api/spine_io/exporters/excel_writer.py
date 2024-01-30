######################################################################################################################
# Copyright (C) 2017-2022 Spine project consortium
# Copyright Spine Database API contributors
# This file is part of Spine Database API.
# Spine Database API is free software: you can redistribute it and/or modify it under the terms of the GNU Lesser
# General Public License as published by the Free Software Foundation, either version 3 of the License, or (at your
# option) any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
# without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Lesser General
# Public License for more details. You should have received a copy of the GNU Lesser General Public License along with
# this program. If not, see <http://www.gnu.org/licenses/>.
######################################################################################################################

"""
A writer for exporting Spine databases to Excel files.

"""
from pathlib import Path
import re
import numpy
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from .writer import Writer, WriterException


class ExcelWriter(Writer):
    def __init__(self, file_path):
        """
        Args:
            file_path (str): path to output file
        """
        super().__init__()
        self._file_path = file_path
        self._workbook = None
        self._current_sheet = None
        self._removable_sheet_names = set()
        self._next_table_name = None
        self._default_sheet_title = None

    def finish(self):
        """See base class."""
        if self._workbook is None:
            return
        for name in self._removable_sheet_names:
            self._workbook.remove(self._workbook[name])
        self._removable_sheet_names.clear()
        if not self._workbook.worksheets:
            self._workbook.create_sheet("Sheet1")
        self._workbook.save(self._file_path)
        self._workbook.close()
        self._workbook = None

    def finish_table(self):
        """See base class."""
        self._current_sheet = None

    def start(self):
        """See base class."""
        if Path(self._file_path).exists():
            try:
                self._workbook = load_workbook(self._file_path)
            except InvalidFileException as e:
                raise WriterException(f"Cannot open Excel file: {e}")
        else:
            self._workbook = Workbook()
            if not self._removable_sheet_names:
                self._removable_sheet_names = set(self._workbook.sheetnames)

    def start_table(self, table_name, title_key):
        """See base class."""
        self._next_table_name = re.sub(INVALID_TITLE_REGEX, "", table_name) if table_name is not None else table_name
        return True

    def _set_current_sheet(self):
        """Gets an existing sheet from workbook or creates a new one if needed."""
        if self._next_table_name is not None:
            if self._next_table_name in self._workbook:
                self._current_sheet = self._workbook[self._next_table_name]
            else:
                self._current_sheet = self._workbook.create_sheet(self._next_table_name)
        else:
            if self._default_sheet_title:
                self._current_sheet = self._workbook[self._default_sheet_title]
            else:
                self._current_sheet = self._workbook.create_sheet(None)
                self._default_sheet_title = self._current_sheet.title
        self._removable_sheet_names.discard(self._current_sheet.title)

    def write_row(self, row):
        """See base class."""
        if self._current_sheet is None:
            self._set_current_sheet()
        row = [_convert_to_excel(cell) for cell in row]
        self._current_sheet.append(row)
        return True


def _convert_to_excel(x):
    """
    Converts parameter values to formats that are comprehensible to openpyxl.

    Args:
        x (Any): a parameter value

    Returns:
        float or str: Excel compatible value
    """
    if isinstance(x, numpy.float_):
        if numpy.isnan(x):
            return "nan"
        return float(x)
    if isinstance(x, numpy.int_):
        return int(x)
    if not isinstance(x, (float, int, str)) and x is not None:
        return str(x)
    return x
