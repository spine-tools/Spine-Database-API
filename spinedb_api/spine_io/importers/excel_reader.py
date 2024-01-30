######################################################################################################################
# Copyright (C) 2017-2022 Spine project consortium
# Copyright Spine Database API contributors
# This file is part of Spine Database API.
# Spine Database API is free software: you can redistribute it and/or modify it under the terms of the GNU Lesser
# General Public License as published by the Free Software Foundation, either version 3 of the License, or (at your
# option) any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
# without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Lesser General
# Public License for more details. You should have received a copy of the GNU Lesser General Public License along with
# this program. If not, see <http://www.gnu.org/licenses/>.
######################################################################################################################

"""
Contains ExcelConnector class and a help function.

"""

from itertools import islice, takewhile, chain
import io
from openpyxl import load_workbook
from .reader import SourceConnection


class ExcelConnector(SourceConnection):
    """Template class to read data from another QThread."""

    # name of data source, ex: "Text/CSV"
    DISPLAY_NAME = "Excel"

    # dict with option specification for source.
    OPTIONS = {
        "header": {"type": bool, "label": "Has header", "default": False},
        "row": {"type": int, "label": "Skip rows", "Minimum": 0, "default": 0},
        "column": {"type": int, "label": "Skip columns", "Minimum": 0, "default": 0},
        "read_until_col": {"type": bool, "label": "Read until empty column on first row", "default": False},
        "read_until_row": {"type": bool, "label": "Read until empty row on first column", "default": False},
    }

    FILE_EXTENSIONS = "*.xlsx;;*.xlsm;;*.xltx;;*.xltm"

    def __init__(self, settings):
        super().__init__(settings)
        self._filename = None
        self._wb = None

    def connect_to_source(self, source, **extras):
        """Connects to Excel file.

        Args:
            source (str): path to file
            **extras: ignored
        """
        if source:
            self._filename = source
            # NOTE: there seems to be no way of closing the workbook
            # when read_only=True, read file into memory first and then
            # open to avoid locking file while toolbox is running.
            with open(self._filename, "rb") as bin_file:
                in_mem_file = io.BytesIO(bin_file.read())
            self._wb = load_workbook(in_mem_file, read_only=True, data_only=True)

    def disconnect(self):
        """Disconnect from connected source."""
        if self._wb:
            self._wb.close()
            self._wb = None
        self._filename = None

    def create_default_mapping(self):
        """See base class"""
        default_mapping = {}
        table_mappings = default_mapping["table_mappings"] = {}
        table_options = default_mapping["table_options"] = {}
        selected_tables = default_mapping["selected_tables"] = []
        for sheet in self.get_tables():
            map_dict, option = create_mapping_from_sheet(self._wb[sheet])
            if map_dict:
                table_mappings[sheet] = [map_dict]
                selected_tables.append(sheet)
            if option:
                table_options[sheet] = option
        return default_mapping

    def get_tables(self):
        """Method that should return Excel sheets as mappings and their options.

        Returns:
            dict: Sheets as mappings and options for each sheet or an empty dictionary if no workbook.

        Raises:
            Exception: If something goes wrong.
        """
        if not self._wb:
            return {}
        try:
            return self._wb.sheetnames
        except Exception as error:
            raise error

    def get_data_iterator(self, table, options, max_rows=-1):
        """
        Return data read from data source table in table. If max_rows is
        specified only that number of rows.
        """
        if not self._wb:
            return iter([]), []
        if table not in self._wb:
            # table not found
            return iter([]), []
        worksheet = self._wb[table]
        # get options
        has_header = options.get("header", False)
        skip_rows = options.get("row", 0)
        skip_columns = options.get("column", 0)
        stop_at_empty_col = options.get("read_until_col", False)
        stop_at_empty_row = options.get("read_until_row", False)
        if max_rows == -1:
            max_rows = None
        else:
            max_rows += skip_rows
        rows = islice(worksheet.iter_rows(), skip_rows, max_rows)
        first_row = next(rows, None)
        if first_row is None:
            return iter([]), []
        read_to_col = None
        if stop_at_empty_col:
            for i, column in enumerate(islice(first_row, skip_columns, None)):
                if column.value is None:
                    read_to_col = i + skip_columns
                    break
        if has_header:
            header = [c.value for c in islice(first_row, skip_columns, read_to_col)]
        else:
            header = []
            rows = chain((first_row,), rows)
        # iterator for selected columns and skipped rows
        data_iterator = (list(cell.value for cell in islice(row, skip_columns, read_to_col)) for row in rows)
        if stop_at_empty_row:
            # add condition to iterator
            data_iterator = takewhile(any, data_iterator)
        return data_iterator, header


def get_mapped_data_from_xlsx(filepath):
    """Returns mapped data from given Excel file assuming it has the default Spine Excel format.

    Args:
        filepath (str): path to Excel file
    """
    connector = ExcelConnector(None)
    connector.connect_to_source(filepath)
    mapping = connector.create_default_mapping()
    table_mappings = {
        name: m for name, m in mapping.get("table_mappings", {}).items() if name in mapping["selected_tables"]
    }
    table_options = {
        name: options
        for name, options in mapping.get("table_options", {}).items()
        if name in mapping["selected_tables"]
    }
    table_types = table_row_types = dict.fromkeys(mapping["selected_tables"], {})
    mapped_data, errors = connector.get_mapped_data(
        table_mappings, table_options, table_types, {}, table_row_types, max_rows=-1
    )
    connector.disconnect()
    return mapped_data, errors


def create_mapping_from_sheet(ws):
    """
    Checks if sheet has the default Spine Excel format, if so creates a
    mapping object for each sheet.
    """
    metadata = _get_metadata(ws)
    if not metadata or "sheet_type" not in metadata:
        return None, None
    header_row = len(metadata) + 2
    if metadata["sheet_type"] == "entity":
        index_dim_count = metadata.get("index_dim_count")
        if index_dim_count is not None:
            index_dim_count = int(index_dim_count)
        header = _get_header(ws, header_row, index_dim_count)
        if not header:
            return None, None
        return _create_entity_mappings(metadata, header, index_dim_count)
    options = {"header": True, "row": len(metadata) + 1, "read_until_col": True, "read_until_row": True}
    if metadata["sheet_type"] == "object_group":
        map_dict = {"map_type": "ObjectGroup", "name": metadata["class_name"], "groups": 0, "members": 1}
        return map_dict, options
    if metadata["sheet_type"] == "alternative":
        map_dict = {"map_type": "Alternative", "name": 0}
        return map_dict, options
    if metadata["sheet_type"] == "scenario":
        map_dict = {"map_type": "Scenario", "name": 0, "active": 1}
        return map_dict, options
    if metadata["sheet_type"] == "scenario_alternative":
        map_dict = {
            "map_type": "ScenarioAlternative",
            "name": 0,
            "scenario_name": 0,
            "alternative_name": 1,
            "before_alternative_name": 2,
        }
        return map_dict, options
    return None, None


def _get_metadata(ws):
    metadata = {}
    for key, value in ws.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
        if not key:
            break
        metadata[key] = value
    return metadata


def _get_header(ws, header_row, index_dim_count):
    if index_dim_count:
        # Vertical header
        header = []
        header_column = index_dim_count
        row = header_row
        while True:
            label = ws.cell(row=row, column=header_column).value
            if label == "index":
                break
            header.append(label)
            row += 1
        return header
    # Horizontal
    header = []
    column = 1
    while True:
        label = ws.cell(row=header_row, column=column).value
        if not label:
            break
        header.append(label)
        column += 1
    return header


def _create_entity_mappings(metadata, header, index_dim_count):
    class_name = metadata["class_name"]
    entity_type = metadata["entity_type"]
    map_dict = {"name": class_name}
    ent_alt_map_type = "row" if index_dim_count else "column"
    if entity_type == "object":
        map_dict["map_type"] = "ObjectClass"
        map_dict["objects"] = {"map_type": ent_alt_map_type, "reference": 0}
    elif entity_type == "relationship":
        entity_dim_count = int(metadata["entity_dim_count"])
        map_dict["map_type"] = "RelationshipClass"
        map_dict["object_classes"] = header[:entity_dim_count]
        map_dict["objects"] = [{"map_type": ent_alt_map_type, "reference": i} for i in range(entity_dim_count)]
    else:
        return None, None
    value_type = metadata.get("value_type")
    if value_type is not None:
        value = {"value_type": value_type}
        if index_dim_count:
            value["extra_dimensions"] = list(range(index_dim_count))
        p_ref = len(header) if index_dim_count else -1
        map_dict["parameters"] = {
            "map_type": "ParameterValue",
            "name": {"map_type": "row", "reference": p_ref},
            "value": value,
            "alternative_name": {"map_type": ent_alt_map_type, "reference": header.index("alternative")},
        }
    options = {
        "header": not index_dim_count,
        "row": len(metadata) + 1,
        "read_until_col": not index_dim_count,
        "read_until_row": not index_dim_count,
    }
    return map_dict, options
